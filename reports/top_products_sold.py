import logging.config
import os
import time
from datetime import datetime  # For working with dates

import matplotlib.pyplot as plt
import pandas as pd  # For working with DataFrames
import seaborn as sns
from PIL import Image
from dotenv import load_dotenv
from matplotlib.ticker import FuncFormatter
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine

# Load logging configuration from file
logging.config.fileConfig('config/logging_config.ini')

logo_path = r'D:\Projects\BOT\reports\trash_media\logo.png'
logo = Image.open(logo_path)
size = 0.4
# Resize the logo to 43% of its original size
resized_width = int(logo.width * size)
resized_height = int(logo.height * size)
logo = logo.resize((resized_width, resized_height))

##################### LOADING IMPORTANT DATA ######################
# Load environment variables from the .env file
env_file_path = r'D:/Projects/.env'
load_dotenv(env_file_path)
# Load data from different sheets in 'promotion.xlsx' into DataFrames
promotion_path = r'D:\Projects\promotion.xlsx'
region_df = pd.read_excel(promotion_path, sheet_name='Region')
types_df = pd.read_excel(promotion_path, sheet_name='TYPES')
##################### ACCESS ENV VARIABLES ######################
db_server = os.getenv("DB_SERVER")
db_database = os.getenv("DB_DATABASE_SERGELI")
db_user = os.getenv("DB_USER")
db_password = os.getenv("DB_PASSWORD")
db_port = os.getenv("DB_PORT")
db_driver_name = os.getenv("DB_DRIVER_NAME")
##################### PROCEDURE NAME ######################
procedure_name = 'zAdmReportDFS_short'  # THIS IS HOURLY DATA GATHERING


# Function to format large numbers
def format_large_numbers(value, pos):
    if value >= 1e9:  # If the value is in billions
        return f'{value / 1e9:.1f}B'
    elif value >= 1e6:  # If the value is in millions
        return f'{value / 1e6:.1f}M'
    elif value >= 1e3:  # If the value is in thousands
        return f'{value / 1e3:.1f}K'
    else:  # For values less than 1000
        return str(int(value))


# Function to abbreviate or truncate the Good names
def abbreviate_good_name(name, max_words=3):
    words = name.split()
    if len(words) <= max_words:
        return name
    else:
        return ' '.join(words[:max_words]) + '...'


def create_bar_plot(df, picture_path, type_value, title, x_column, y_column):
    try:
        logging.info(f'Starting to plot {title} for {picture_path}')

        # Select the top 20 rows for plotting
        top_rows = 20
        df = df.head(top_rows).copy()

        # Abbreviate good names for better display in the plot
        df[y_column] = df[y_column].apply(lambda x: abbreviate_good_name(x, max_words=3))

        # Select color palette based on type_value
        if type_value == 'ROZ':
            palette = sns.color_palette("tab20b")
        elif type_value == 'Сеть':
            palette = sns.color_palette("tab20c")
        elif type_value == 'Опт':
            palette = sns.color_palette("tab20")
        else:
            palette = sns.color_palette("tab20")  # Default palette

        output_picture_path = f'reports/trash_media/Top_20_Goods_{picture_path.split("_")[1]}_{type_value}.png'
        logging.info(f'Picture path is: {output_picture_path}')

        # Create a new figure
        fig, ax = plt.subplots(figsize=(18, 10))

        # Plotting horizontal bars using Seaborn
        bars = sns.barplot(x=x_column, y=y_column, data=df, ax=ax, palette=palette, errorbar=None)
        logging.info(f'Plot for {output_picture_path} is ready.')

        # Adding labels and title
        ax.set_xlabel(f'{x_column}')
        ax.set_ylabel(f'{y_column}')
        ax.set_title(f'{title} : {type_value}')

        # Annotating each bar with its value
        for bar, value in zip(bars.patches, df[x_column]):
            ax.text(bar.get_width() + 0.005 * bar.get_width(), bar.get_y() + bar.get_height() / 2,
                    format_large_numbers(value, None), va='center')

        # Format x-axis labels using the formatter function
        ax.xaxis.set_major_formatter(FuncFormatter(format_large_numbers))

        # Adding the logo to the figure
        fig.figimage(logo, xo=1051.7, yo=65, alpha=0.5)  # Adjust xo, yo, and alpha as needed

        # Save the plot to a file
        plt.savefig(output_picture_path, bbox_inches='tight')
        plt.close()
        logging.info(f'Successfully saved plot to {output_picture_path}')

    except Exception as e:
        logging.error(f"Error in create_bar_plot: {e}")
        plt.close()


def create_work_sheet(workbook, type_value, pivot_table, i):
    # Create a new sheet with the type name
    worksheet = workbook.create_sheet(title=str(type_value))
    colors_for_sheet = ['FFC4C4', 'F3B95F', 'EAFFD0']

    worksheet.sheet_properties.tabColor = colors_for_sheet[i - 1]
    # Create a new named style for the header
    header_style = NamedStyle(name=f'header_style_{i}',
                              fill=PatternFill(start_color=colors_for_sheet[i - 1],
                                               end_color=colors_for_sheet[i - 1], fill_type='solid'))

    # Add pivot table to the sheet
    for row_idx, (index, values) in enumerate(pivot_table.iterrows(), start=2):
        worksheet.cell(row=row_idx, column=1, value=index)

        # Iterate through the values and add them to the respective columns
        for col_idx, (col, value) in enumerate(values.items(), start=2):
            worksheet.cell(row=row_idx, column=col_idx, value=value)

    # Add headers for the pivot table
    for col_idx, col_name in enumerate(pivot_table.columns, start=2):
        worksheet.cell(row=1, column=col_idx, value=col_name)

    # Apply the header style to the first row
    for cell in worksheet[1]:
        cell.style = header_style
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 2. Autofit all columns
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                value = str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)
            except Exception as e:
                logging.error(e)

        adjusted_width = (max_length + 1.5)
        column_letter = get_column_letter(column[0].column)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # 3. Format columns with thousands separators
    number_format = NamedStyle(name=f'number_format_{i}', number_format='### ### ### ##0')

    # Specify the columns to format based on float64 datatype
    float64_columns = pivot_table.select_dtypes(include=['float64']).columns

    for col in float64_columns:
        col_index = pivot_table.columns.get_loc(col) + 2  # 1-based index, starting from column 2
        col_letter = get_column_letter(col_index)

        for cell in worksheet[col_letter][1:]:  # Start from the second row assuming the first row is headers
            try:
                formatted_value = "{:,.2f}".format(float(cell.value))
                cell.value = float(cell.value)
                cell.style = number_format
            except (ValueError, TypeError) as error:
                logging.error(error)

    # Apply background color to all cells in the TOTAL (2nd) column
    for row in worksheet.iter_rows(min_row=1, max_row=16, min_col=1,
                                   max_col=3):
        for cell in row:
            cell.fill = PatternFill(start_color=colors_for_sheet[i - 1], end_color=colors_for_sheet[i - 1],
                                    fill_type='solid')

    # Add borders to all cells with data
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1,
                                   max_col=worksheet.max_column):
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))


def save_workbook(workbook, output_file_path):
    try:
        workbook.save(output_file_path)
    except Exception as e:
        logging.error(f"Error saving workbook: {e}")


def top_product_sold_generator(start_date, end_date, current_month: bool = True, ) -> dict:
    logging.info('Started running top_product_sold_generator')

    start_date = datetime.strptime(start_date, '%Y%m%d')
    end_date = datetime.strptime(end_date, '%Y%m%d')
    logging.info(f'Start date: {start_date}, End date: {end_date}')

    # Construct the connection string
    conn_str = f"mssql+pyodbc://{db_user}:{db_password}@{db_server}:{db_port}/{db_database}?driver={db_driver_name}"

    engine = create_engine(conn_str)

    YEAR = start_date.strftime('%Y')
    START_MONTH = start_date.strftime('%m')

    END_YEAR = end_date.strftime('%Y')
    END_MONTH = end_date.strftime('%m')
    sql_query = f"""
            DECLARE @DateBegin DATE = ?;
            DECLARE @DateEnd DATE = ?;

            EXEC {procedure_name}
                @DateBegin = @DateBegin,
                @DateEnd = @DateBegin;
        """

    #####################  EXECUTION  ######################
    df = pd.read_sql_query(sql_query, engine, params=(start_date, end_date))
    df.columns = ['DocumentType', 'Invoice Number', 'Goodid', 'Good', 'Manufacturer', 'inn', 'ClientName',
                  'InvoiceManager', 'ClientMan', 'PaymentTerm', 'BasePrice', 'SellingPrice', 'Quantity', 'DateEntered',
                  'BaseAmount', 'TotalAmount']

    ##################### BASIC FILTER ######################
    df['DateEntered'] = pd.to_datetime(df['DateEntered'])
    df = df[
        (df['DateEntered'].dt.year == int(YEAR)) &
        (df['DateEntered'].dt.month >= int(START_MONTH)) &
        (df['DateEntered'].dt.month <= int(END_MONTH)) &
        df['DocumentType'].isin(
            ['Оптовая реализация',
             'Финансовая скидка'
             ])
        & ~df['InvoiceManager'].isin(['Администратор', 'Джалилов Шавкат', 'Бочкарева Альвина'])
        ]
    region_df['ClientMan'] = region_df['ClientMan'].str.title()
    df['ClientMan'] = df['ClientMan'].str.title()

    logging.info(f'{top_product_sold_generator.__name__} : Finished Filtering DataFrame: {df.shape}')
    df = pd.merge(df, region_df[['ClientMan', 'Region']], left_on='ClientMan', right_on='ClientMan', how='left')
    df['inn_temp'] = pd.to_numeric(df['inn'], errors='coerce')
    types_df['INN_temp'] = pd.to_numeric(types_df['INN'], errors='coerce')
    df = pd.merge(df, types_df[['INN_temp', 'TYPE', 'RegionType']], left_on='inn_temp', right_on='INN_temp', how='left')
    logging.info('Performing advanced filtering and formatting...')
    df['OXVAT'] = df['inn'].map(df['inn'].value_counts())

    df.loc[df['TYPE'] == 'ROZ', 'RegionType'] = df['Region']
    df['TYPE'] = df['TYPE'].fillna('ROZ')
    df.loc[df['TYPE'] == 'ROZ', 'RegionType'] = df['Region']
    df.drop(['INN_temp', 'inn_temp'], axis=1, inplace=True)
    df = df[df['RegionType'] != 'Админ']
    logging.info(f'{top_product_sold_generator.__name__} : DataFrame successfully created: {df.shape}')
    top_files = {
        'top_revenue_products': {
            'file_name': f'TOP_REVENUE_PRODUCTS_SOLD-{start_date.strftime("%Y%m%d")}-{end_date.strftime("%Y%m%d")}.xlsx',
            'description': 'Товары, приносящие наибольший доход, среди клиентов в различных регионах и типах на',
            'picture_name': 'revenue',
        },
        'high_volume_products': {
            'file_name': f'HIGH_VOLUME_PRODUCTS-{start_date.strftime("%Y%m%d")}-{end_date.strftime("%Y%m%d")}.xlsx',
            'description': 'Товары с наибольшим объемом(колич) продаж среди клиентов в различных регионах и типах на',
            'picture_name': 'volume',
        },
        'client_fav_products': {
            'file_name': f'CLIENT_FAVORITE_PRODUCTS-{start_date.strftime("%Y%m%d")}-{end_date.strftime("%Y%m%d")}.xlsx',
            'description': 'Товары, популярные среди клиентов в различных регионах и разных типов на',
            'picture_name': 'favorite',

        }
    }
    top_revenue_products(df=df, output_file_path=top_files['top_revenue_products']['file_name'],
                         picture_path=top_files['top_revenue_products']['picture_name'])
    high_volume_products(df=df, output_file_path=top_files['high_volume_products']['file_name'],
                         picture_path=top_files['high_volume_products']['picture_name'])
    client_fav_products(df=df, output_file_path=top_files['client_fav_products']['file_name'],
                        picture_path=top_files['client_fav_products']['picture_name'])

    return top_files


def top_revenue_products(df, output_file_path, picture_path):
    logging.info(f'{top_revenue_products.__name__}: Creating worksheet {output_file_path}, {df.shape}')

    # Record the start time
    start_time = time.time()

    # Create a new workbook and remove the default sheet
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    # Set Seaborn palette
    pastel_palette = sns.color_palette("tab20b")

    # Iterate over unique types in the dataframe
    for i, type_value in enumerate(sorted(df['TYPE'].unique()), start=1):
        # Filter dataframe for the current type
        type_df = df[df['TYPE'] == type_value]
        logging.info(f"Processing sheet: {type_value}, DataFrame shape: {type_df.shape}")

        # Group by Good and Region, then sum the TotalAmount
        grouped_df = type_df.groupby(['Good', 'RegionType', 'ClientMan'], observed=False).agg(
            {'TotalAmount': 'sum'}).reset_index()
        logging.info(f"Grouped DataFrame shape: {grouped_df.shape}")

        # Create a pivot table
        pivot_table = pd.pivot_table(grouped_df, values='TotalAmount', index=['Good'], columns=['RegionType'],
                                     aggfunc='sum', fill_value=0)

        # Calculate Grand Total
        pivot_table['TOTAL'] = pivot_table.sum(axis=1)
        # Sort by 'TOTAL' in descending order
        pivot_table.sort_values(by='TOTAL', ascending=False, inplace=True)

        # Reorder the columns
        pivot_table = pivot_table[['TOTAL'] + list(pivot_table.columns[:-1])]

        # Exclude columns with total 0 in the pivot_table
        pivot_table = pivot_table.loc[:, (pivot_table != 0).any(axis=0)]
        pivot_table = pivot_table.loc[(pivot_table != 0).any(axis=1)]
        pivot_table.reset_index(drop=False, inplace=True)
        pivot_table.index += 1
        pivot_table.index.name = '#'
        logging.info(f"Sheet {i} name: {type_value}, Pivot tabble shape: {pivot_table.shape}")
        create_work_sheet(workbook=workbook, type_value=type_value, pivot_table=pivot_table, i=i)

        # PLOTTING THE BARS
        # create_bar_plot(df=pivot_table,
        #                 title='20 TOP REVENUE PRODUCTS',
        #                 type_value=type_value,
        #                 x_column='TOTAL',
        #                 y_column='Good',
        #                 picture_path=picture_path
        #                 # hue_column='RegionType'
        #                 )

    save_workbook(workbook=workbook, output_file_path=output_file_path)

    # Log the elapsed time
    elapsed_time = time.time() - start_time
    logging.info(f"Elapsed time: {elapsed_time:.2f} seconds")


def client_fav_products(df, output_file_path, picture_path):
    logging.info(f'Starting process to create {output_file_path}.')

    # Record the start time
    start_time = time.time()

    # Create a new workbook and remove the default sheet
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    # Set Seaborn palette
    pastel_palette = sns.color_palette("tab20")

    # Iterate over unique types in the dataframe
    for i, type_value in enumerate(sorted(df['TYPE'].unique()), start=1):
        try:
            # Filter dataframe for the current type
            type_df = df[df['TYPE'] == type_value]
            logging.info(f"Processing sheet: {type_value}, DataFrame shape: {type_df.shape}")

            # Group by Good and Region, then count the unique 'inn'
            grouped_df = type_df.groupby(['Good', 'RegionType'], observed=False).agg({'inn': 'nunique'}).reset_index()
            logging.info(f"Grouped DataFrame shape: {grouped_df.shape}")

            # Create a pivot table
            pivot_table = pd.pivot_table(grouped_df, index='Good', columns='RegionType', values='inn', aggfunc='sum',
                                         fill_value=0)
            logging.info(f"Pivot Table shape: {pivot_table.shape}")

            # Drop the 'Admin' column
            pivot_table.drop(columns=['Админ'], inplace=True, errors='ignore')

            # Calculate Grand Total
            pivot_table['TOTAL CLIENTS'] = pivot_table.sum(axis=1)

            # Sort by 'TOTAL CLIENTS' in descending order
            pivot_table.sort_values(by='TOTAL CLIENTS', ascending=False, inplace=True)

            # Reorder the columns
            pivot_table = pivot_table[['TOTAL CLIENTS'] + list(pivot_table.columns[:-1])]

            # Exclude columns with total 0 in the pivot_table
            pivot_table = pivot_table.loc[:, (pivot_table != 0).any(axis=0)]
            pivot_table = pivot_table.loc[(pivot_table != 0).any(axis=1)]
            pivot_table.reset_index(drop=False, inplace=True)
            pivot_table.index += 1
            pivot_table.index.name = '#'
            logging.info(f"Sheet {i} name: {type_value}")

            # PLOTTING THE BARS
            # Select the top 20 rows
            top_20_rows = pivot_table.head(20).copy()
            top_20_rows['Abbreviated Good'] = top_20_rows['Good'].apply(lambda x: abbreviate_good_name(x, max_words=3))

            # create_bar_plot(df=top_20_rows,
            #                 title='20 TOP CLIENTS FAVORITE PRODUCTS',
            #                 type_value=type_value,
            #                 x_column='TOTAL CLIENTS',
            #                 y_column='Abbreviated Good',
            #                 # hue_column='RegionType',
            #                 picture_path=picture_path
            #                 )

            # Create a new sheet with the type name
            worksheet = workbook.create_sheet(title=str(type_value)[:30])  # Limit title to 31 characters
            colors_for_sheet = ['FFC4C4', 'F3B95F', 'EAFFD0']
            worksheet.sheet_properties.tabColor = colors_for_sheet[i % len(colors_for_sheet)]

            # Create a new named style for the header
            header_style = NamedStyle(name=f'header_style_{i}',
                                      fill=PatternFill(start_color=colors_for_sheet[i % len(colors_for_sheet)],
                                                       end_color=colors_for_sheet[i % len(colors_for_sheet)],
                                                       fill_type='solid'))

            # Add pivot table to the sheet
            for row_idx, (index, values) in enumerate(pivot_table.iterrows(), start=2):
                worksheet.cell(row=row_idx, column=1, value=index)
                for col_idx, (col, value) in enumerate(values.items(), start=2):
                    worksheet.cell(row=row_idx, column=col_idx, value=value)

            # Add headers for the pivot table
            for col_idx, col_name in enumerate(pivot_table.columns, start=2):
                worksheet.cell(row=1, column=col_idx, value=col_name)

            # Apply the header style to the first row
            for cell in worksheet[1]:
                cell.style = header_style
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Autofit all columns
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        value = str(cell.value)
                        if len(value) > max_length:
                            max_length = len(value)
                    except:
                        pass
                adjusted_width = (max_length + 1.8)
                column_letter = get_column_letter(column[0].column)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Format columns with thousands separators
            number_format = NamedStyle(name=f'number_format_{i}', number_format='### ### ### ##0')
            float64_columns = pivot_table.select_dtypes(include=['float64']).columns

            for col in float64_columns:
                col_index = pivot_table.columns.get_loc(col) + 2
                col_letter = get_column_letter(col_index)

                for cell in worksheet[col_letter][1:]:
                    try:
                        cell.value = float(cell.value)
                        cell.style = number_format
                    except (ValueError, TypeError) as error:
                        logging.warning(f"Error formatting cell value: {error}")

            # Apply background color to all cells in the TOTAL (2nd) column
            for row in worksheet.iter_rows(min_row=1, max_row=26, min_col=1, max_col=3):
                for cell in row:
                    cell.fill = PatternFill(start_color=colors_for_sheet[i % len(colors_for_sheet)],
                                            end_color=colors_for_sheet[i % len(colors_for_sheet)], fill_type='solid')

            # Add borders to all cells with data
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1,
                                           max_col=worksheet.max_column):
                for cell in row:
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                         bottom=Side(style='thin'))
        except Exception as e:
            logging.error(f"Error processing type {type_value}: {e}")

    try:
        # Save the workbook
        workbook.save(output_file_path)
        end_time = time.time()
        elapsed_time = round(end_time - start_time, 2)
        logging.info(f"Workbook saved successfully to {output_file_path}. Process took: {elapsed_time} seconds.")
    except Exception as e:
        logging.error(f"Error saving workbook: {e}")


def high_volume_products(df, output_file_path, picture_path):
    logging.info("Creating High Volume and Exporting🎨....")
    # Record the start time
    start_time = time.time()

    # Create a new workbook and remove the default sheet
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    pastel_palette = sns.color_palette("tab20c")

    # Iterate over unique types in the dataframe
    for i, type_value in enumerate(sorted(df['TYPE'].unique()), start=1):
        # Filter dataframe for the current type
        type_df = df[df['TYPE'] == type_value]
        logging.info(f"Processing sheet: {type_value}, DataFrame shape: {type_df.shape}")

        # Assuming df is the DataFrame obtained from the SQL query
        result_df = type_df.copy()
        logging.info(f"Result DataFrame shape: {result_df.shape}")

        # Group by Good and Region, then count the unique 'inn'
        grouped_df = result_df.groupby(['Good', 'RegionType'], observed=False).agg({'Quantity': 'sum'}).reset_index()
        logging.info(f"Grouped DataFrame shape: {grouped_df.shape}")

        # Create a pivot table
        pivot_table = pd.pivot_table(grouped_df, index='Good', columns='RegionType', values='Quantity', aggfunc='sum',
                                     fill_value=0, observed=False)

        # Drop the 'Admin' column
        pivot_table.drop(columns=['Админ'], inplace=True, errors='ignore')

        # Calculate Grand Total
        pivot_table['TOTAL QUANTITY'] = pivot_table.sum(axis=1)

        # Sort by 'TOTAL' in descending order
        pivot_table.sort_values(by='TOTAL QUANTITY', ascending=False, inplace=True)

        # Reorder the columns
        pivot_table = pivot_table[['TOTAL QUANTITY'] + list(pivot_table.columns[:-1])]

        # Exclude columns with total 0 in the pivot_table
        pivot_table = pivot_table.loc[:, (pivot_table != 0).any(axis=0)]
        pivot_table = pivot_table.loc[(pivot_table != 0).any(axis=1)]
        pivot_table.reset_index(drop=False, inplace=True)
        pivot_table.index += 1
        pivot_table.index.name = '#'
        print(f"Sheet {i} name: {type_value}")

        # create_bar_plot(df=pivot_table,
        #                 title='20 TOP HIGH VOLUME PRODUCTS',
        #                 type_value=type_value,
        #                 x_column='TOTAL QUANTITY',
        #                 y_column='Good',
        #                 picture_path=picture_path
        #                 # hue_column='RegionType'
        #                 )
        create_work_sheet(workbook=workbook, type_value=type_value, pivot_table=pivot_table, i=i)

    save_workbook(workbook=workbook, output_file_path=output_file_path)

    logging.info(f"{high_volume_products.__name__} is completed in {time.time() - start_time} seconds.")
