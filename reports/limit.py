# -*- coding: UTF-8 -*-

import os
import time
from datetime import datetime  # For working with dates

import pandas as pd  # For working with DataFrames
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine  # For creating a connection engine


def limit_generator():
    print('Started running Limit')
    start_time = time.time()
    # Load environment variables from the .env file
    env_file_path = 'D:/Projects/.env'
    load_dotenv(env_file_path)

    # Get today's date
    today_date = datetime.now().strftime('%d %b')
    output_file_path = f'LIMIT - {today_date}.xlsx'
    # Load data from different sheets in 'promotion.xlsx' into DataFrames
    promotion_path = 'D:\Projects\promotion.xlsx'
    region_df = pd.read_excel(promotion_path, sheet_name='Region')
    # Access the environment variables
    db_server = os.getenv("DB_SERVER")
    db_database = os.getenv("DB_DATABASE_ASKGLOBAL")
    db_user = os.getenv("DB_USER")
    db_password = os.getenv("DB_PASSWORD")
    db_port = os.getenv("DB_PORT")
    db_driver_name = os.getenv("DB_DRIVER_NAME")

    # Construct the connection string
    conn_str = f"mssql+pyodbc://{db_user}:{db_password}@{db_server}:{db_port}/{db_database}?driver={db_driver_name}"

    procedure_name = "atCalcExtraLimitsByInn"
    engine = create_engine(conn_str)

    # Build the SQL query with parameterized query
    sql_query = f""" EXEC {procedure_name} """

    # 💀💀💀 EXECUTION!!! ⚠️⚠️⚠️
    df = pd.read_sql_query(sql_query, engine)

    # Making it similiar case title so the data loss is prevented
    region_df['ClientMan'] = region_df['ClientMan'].str.title()
    df['ClientMan'] = df['Менеджер/ка'].str.strip()
    df['ClientMan'] = df['ClientMan'].str.title()

    # Merge with 'region_df' DataFrame based on 'ClientMan'
    df = pd.merge(df, region_df[['ClientMan', 'Region']], left_on='ClientMan', right_on='ClientMan', how='left')

    df = df[['Филиал', 'Region', 'ClientMan', 'ИНН клиента', 'Наименование клиента', 'Лимит клиента', 'Дебиторка',
             'Свободный лимит']]
    df = df[df['Свободный лимит'] > 200_000]
    df.sort_values(by=['Region', 'ClientMan'], inplace=True)
    df.rename(columns={'Region': 'Регион', 'ClientMan': "Ответственный Менеджер"}, inplace=True)
    end_time = time.time()
    df.to_excel(output_file_path, index=False)
    print(f"Data Preparation took: {round(end_time - start_time, 0)} seconds.")
    print(f"Its time to format💅🏻")
    # Load the existing workbook
    workbook = load_workbook(output_file_path)

    # Access the default sheet (assuming it's the only sheet in the workbook)
    worksheet = workbook.active

    # 1. Color first row (headers) with 4CB9E7 color code
    header_style = NamedStyle(name='header_style',
                              fill=PatternFill(start_color='4CB9E7', end_color='4CB9E7', fill_type='solid'))

    for cell in worksheet[1]:
        cell.style = header_style

    # 2. Autofit all columns
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                value = str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)
            except:
                pass
        adjusted_width = (max_length + 5)
        column_letter = get_column_letter(column[0].column)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # 3. Format columns with thousands separators
    number_format = NamedStyle(name='number_format', number_format='### ### ### ##0')

    # Specify the columns to format based on sorted_pivoted_df columns
    for i, column in enumerate(worksheet.iter_cols(min_col=5, max_col=worksheet.max_column), start=2):
        for cell in column[1:]:
            try:
                # Convert the value to a float, round it, and then format with thousands separators
                formatted_value = "{:,.0f}".format(round(float(cell.value)))
                cell.value = int(float(cell.value))
                cell.style = number_format
            except (ValueError, TypeError):
                # Handle cases where the cell value is not a valid number
                pass

            try:
                # Convert the value to a float, round it, and then format with thousands separators
                formatted_value = "{:,.0f}".format(round(float(cell.value)))
                cell.value = int(float(cell.value))
                cell.style = number_format
            except (ValueError, TypeError):
                # Handle cases where the cell value is not a valid number
                pass
    # Get the last column letter
    last_column_letter = get_column_letter(worksheet.max_column)

    # Apply background color to all cells in the last column
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=worksheet.max_column,
                                   max_col=worksheet.max_column):
        for cell in row:
            cell.fill = PatternFill(start_color='A0D8B3', end_color='A0D8B3', fill_type='solid')

    # Add borders to all cells with data
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))

    workbook.save(output_file_path)
