import os
import pandas as pd
from sqlalchemy import create_engine

from datetime import datetime, timedelta
from dotenv import load_dotenv

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter

env_file_path = 'D:/Projects/.env'
load_dotenv(env_file_path)

dail_sales_path = 'D:/Projects/daily_sales.xlsx'


def planned_actual_sales_generator(output_file_path):
    daily_plan = pd.read_excel(dail_sales_path, sheet_name='clients_list')
    ##################### ACCESS ENV VARIABLES ######################
    db_server = os.getenv("DB_SERVER")
    db_database = os.getenv("DB_DATABASE_SERGELI")
    db_user = os.getenv("DB_USER")
    db_password = os.getenv("DB_PASSWORD")
    db_port = os.getenv("DB_PORT")
    db_driver_name = os.getenv("DB_DRIVER_NAME")

    procedure_name = os.getenv("HOURLY_SHORT")  # THIS IS HOURLY DATA GATHERING
    yesterday_date = (datetime.now() - timedelta(days=1))
    today_date = (datetime.now())
    CURRENT_MONTH = datetime.now().month
    CURRENT_YEAR = datetime.now().year

    # Construct the connection string
    conn_str = f"mssql+pyodbc://{db_user}:{db_password}@{db_server}:{db_port}/{db_database}?driver={db_driver_name}"
    engine = create_engine(conn_str)

    sql_query = f"""
        DECLARE @DateBegin DATE = ?;
        DECLARE @DateEnd DATE = ?;
    
        EXEC {procedure_name}
            @DateBegin = @DateBegin,
            @DateEnd = @DateBegin;
    """

    # Execute the query with yesterday's date
    df = pd.read_sql_query(sql_query, engine, params=(yesterday_date.strftime('%Y%m%d'), today_date.strftime('%Y%m%d')))

    ##################### BASIC FILTER ######################
    df['DataEntered'] = pd.to_datetime(df['DataEntered'])
    df.columns = ['DocKind',
                  'InvoiceNumber',
                  'GoodId',
                  'GoodName',
                  'Manufacturer',
                  'INN',
                  'ClientName',
                  'InvoiceManager',
                  'ClientMan', 'PaymentTerm',
                  'BasePrice', 'SellingPrice',
                  'Quantity', 'DataEntered', 'BaseAmount', 'TotalAmount']
    df = df[(df['DataEntered'].dt.month == CURRENT_MONTH)
            & (df['DataEntered'].dt.year == CURRENT_YEAR)
            & (df['DataEntered'].dt.date == yesterday_date.date())
            & df['DocKind'].isin(['Оптовая реализация', 'Финансовая скидка'])]
    ##################### CONVERT TO CATEGORICAL DATA TYPE ######################
    daily_plan['INN'] = daily_plan['INN'].astype(str)
    df['INN'] = df['INN'].astype(str)
    df.sort_values(by='DataEntered', ascending=False, inplace=True)

    result_pivot = pd.pivot_table(df, values='TotalAmount', index=['INN', 'ClientName'], aggfunc='sum',
                                  fill_value=0).reset_index()

    result = pd.merge(daily_plan, result_pivot, on=['INN', 'ClientName'], how='left').fillna(0)
    result['Difference'] = result['TotalAmount'] - result['SALES PLAN']
    result.sort_values(by='TotalAmount', ascending=False, inplace=True)
    sales_made = result[result['TotalAmount'] > 0].copy()
    no_sales_made = result[result['TotalAmount'] <= 0].copy().sort_values(by='Region')

    type_values = {
        'SALES': {'data_frame': sales_made, 'color': 'FFC4C4'},
        'NO SALES': {'data_frame': no_sales_made, 'color': 'F3B95F'},
        'ALL CLIENTS': {'data_frame': result, 'color': 'EAFFD0'},
    }

    def create_work_sheet(workbook, type_value, pivot_table, worksheet_color) -> None:
        # Create a new sheet with the type name
        worksheet = workbook.create_sheet(title=str(type_value))

        worksheet.sheet_properties.tabColor = worksheet_color

        # Create a new named style for the header (create only once)
        header_style = NamedStyle(name=f'{type_value}_header_style',
                                  fill=PatternFill(start_color=worksheet_color,
                                                   end_color=worksheet_color, fill_type='solid'))

        # Add headers for the pivot table
        header_values = list(pivot_table.columns)
        for col_idx, header_value in enumerate(header_values, start=1):
            worksheet.cell(row=1, column=col_idx, value=header_value)
            worksheet.cell(row=1, column=col_idx).style = header_style
            worksheet.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')

        # Add data to the sheet
        for row_idx, row in enumerate(pivot_table.values, start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)

        # Autofit all columns
        for col_idx, column in enumerate(worksheet.columns, start=1):
            max_length = 0
            for cell in column:
                try:
                    value = str(cell.value)
                    if len(value) > max_length:
                        max_length = len(value)
                except TypeError:
                    pass
            adjusted_width = (max_length + 1.7) * 1.2  # Adjust width for better readability
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = adjusted_width

        # Format columns with thousands separators
        number_format = NamedStyle(name=f'{type_value}_number_format', number_format='### ### ### ##0')

        for col_idx, col_name in enumerate(pivot_table.columns, start=1):
            col_index = col_idx
            col_letter = get_column_letter(col_index)
            if pivot_table[col_name].dtype in ['float64', 'int64']:
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=col_idx,
                                               max_col=col_idx):
                    for cell in row:
                        try:
                            if cell.value is not None:
                                formatted_value = "{:,.2f}".format(float(cell.value))
                                cell.value = float(cell.value)
                                cell.style = number_format
                        except (ValueError, TypeError) as error:
                            print(error)

        # Add borders to all cells with data
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1,
                                       max_col=worksheet.max_column):
            for cell in row:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                     bottom=Side(style='thin'))

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    def save_workbook(workbook: Workbook, output_file_path: str):
        try:
            workbook.save(output_file_path)
        except Exception as e:
            print(f"Error saving workbook: {e}")

    # Create named styles outside the loop
    header_styles = {}
    number_format_styles = {}
    for type_value, value_info in type_values.items():
        color = value_info['color']
        header_style = NamedStyle(name=f'{type_value}_header_style',
                                  fill=PatternFill(start_color=color,
                                                   end_color=color, fill_type='solid'))
        number_format_style = NamedStyle(name=f'{type_value}_number_format', number_format='### ### ### ##0')
        header_styles[type_value] = header_style
        number_format_styles[type_value] = number_format_style

    for type_value, value_info in type_values.items():
        data_frame = value_info['data_frame']
        color = value_info['color']
        create_work_sheet(workbook=workbook, pivot_table=data_frame, type_value=type_value, worksheet_color=color)

    save_workbook(workbook=workbook, output_file_path=output_file_path)
