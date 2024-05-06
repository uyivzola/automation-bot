import os
from datetime import datetime, date, timedelta

import pandas as pd  # For working with DataFrames
from dotenv import load_dotenv
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine  # For crea
from openpyxl.styles import Alignment, Border, Side, NamedStyle, PatternFill
from openpyxl import Workbook

##################### LOADING IMPORTANT DATA ######################
# Load environment variables from the .env file
env_file_path = 'D:/Projects/.env'
load_dotenv(env_file_path)
# Giving output file name
output_file_path = 'ГАТ.xlsx'
# Load data from different sheets in 'promotion.xlsx' into DataFrames
promotion_path = 'D:/Projects/promotion.xlsx'
region_df = pd.read_excel(promotion_path, sheet_name='Region')
aksiya_df = pd.read_excel(promotion_path, sheet_name='Aksiya')
paket_df = pd.read_excel(promotion_path, sheet_name='Paket')
types_df = pd.read_excel(promotion_path, sheet_name='TYPES')
okm_df = pd.read_excel(promotion_path, sheet_name='OKM')
##################### ACCESS ENV VARIABLES ######################
db_server = os.getenv("DB_SERVER")
db_database = os.getenv("DB_DATABASE_SERGELI")
db_user = os.getenv("DB_USER")
db_password = os.getenv("DB_PASSWORD")
db_port = os.getenv("DB_PORT")
db_driver_name = os.getenv("DB_DRIVER_NAME")
##################### PROCEDURE NAME ######################
procedure_name = 'gTOandFSkidka'  # THIS IS HOURLY DATA GATHERING
start_of_current_month = date(date.today().year, date.today().month, 1).strftime('%Y%m%d')
tomorrow_date = (datetime.now() + timedelta(days=1)).strftime('%Y%m%d')

##################### CONNECTION STRING AND SQL QUERY ######################
# Construct the connection string
conn_str = f"mssql+pyodbc://{db_user}:{db_password}@{db_server}:{db_port}/{db_database}?driver={db_driver_name}"
engine = create_engine(conn_str)

sql_query = f"""
DECLARE @DateBegin DATETIME;
DECLARE @DateEnd DATETIME;

-- Calculate the start of the current month
SET @DateBegin = DATEADD(month, DATEDIFF(month, 0, GETDATE()), 0);

-- Set the end date as the current date
SET @DateEnd = EOMONTH(GETDATE())

-- Call the stored procedure with the calculated dates
EXEC gTOandFSkidka @DateBegin, @DateEnd;
"""


def okm_generator():
    #####################  EXECUTION  ######################
    df = pd.read_sql_query(sql_query, engine)
    df.columns = ['DocName', 'InvoiceNumber', 'Year', 'Month', 'Data', 'GoodId', 'Good', 'Producer',
                  'SerialNumber', 'ClientName', 'INN', 'City', 'ClientType', 'InvoiceManager',
                  'ClientMan', 'Store', 'StoreDep', 'FSkidka', 'Comment',
                  'DownPayment', 'PaymentTerm', 'BasePrice', 'InPrice', 'OutPrice',
                  'OutQuantity', 'OutSumma', 'prlPrice', 'TypeSkidka', 'TypeTovSkdikda',
                  'ConnectedInvoice', 'DataEntered']
    df['ClientMan'] = df['ClientMan'].str.lower()
    region_df['ClientMan'] = region_df['ClientMan'].str.lower()
    df_tovar = df[df['TypeTovSkdikda'].isin(['Ацекард 30+10', 'Гинокапс Форте 7+2', 'Тризим 10+1'])]
    good_ids = okm_df.GoodId
    selected_invoices = df[df['InvoiceNumber'].isin(df_tovar['ConnectedInvoice'])]
    # Filter the selected rows based on 'GoodId'
    df_selected = selected_invoices[selected_invoices['GoodId'].isin(good_ids)]
    df_final = pd.concat([df_tovar, df_selected])

    df_final = pd.merge(df_final, region_df[['ClientMan', 'Region']], left_on='ClientMan', right_on='ClientMan',
                        how='left')
    df_final = pd.merge(df_final, okm_df[['GoodId', 'Aksiya']], left_on='GoodId', right_on='GoodId', how='left')

    def save_workbook(workbook, output_file_path):
        try:
            workbook.save(output_file_path)
        except Exception as e:
            print(f"Error saving workbook: {e}")

    def create_work_sheet(workbook, type_value, pivot_table, i):
        # Create a new sheet with the type name
        worksheet = workbook.create_sheet(title=str(type_value))
        colors_for_sheet = ['FFC4C4', 'F3B95F', 'EAFFD0']

        worksheet.sheet_properties.tabColor = colors_for_sheet[i - 1]
        # Create a new named style for the header
        header_style = NamedStyle(name=f'header_style_{i}',
                                  fill=PatternFill(start_color=colors_for_sheet[i - 1],
                                                   end_color=colors_for_sheet[i - 1], fill_type='solid'))
        # Add headers for the pivot table - Write the header of the MultiIndex explicitly
        header_values = list(pivot_table.index.names) + list(pivot_table.columns)

        for col_idx, header_value in enumerate(header_values, start=1):
            worksheet.cell(row=1, column=col_idx, value=header_value)

        # Add pivot table to the sheet
        for row_idx, (index, values) in enumerate(pivot_table.iterrows(), start=2):
            # Populate each level of the MultiIndex into separate columns
            for col_idx, idx_value in enumerate(index, start=1):
                worksheet.cell(row=row_idx, column=col_idx, value=idx_value)

            # Iterate through the values and add them to the respective columns
            for col_idx, (col, value) in enumerate(values.items(), start=len(index) + 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)

        # Add headers for the pivot table
        for col_idx, col_name in enumerate(pivot_table.columns, start=4):
            worksheet.cell(row=1, column=col_idx, value=col_name)

        # Apply the header style to the first row
        for cell in worksheet[1]:
            cell.style = header_style
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 2. Autofit all columns
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    value = str(cell.value)
                    if len(value) > max_length:
                        max_length = len(value)
                except:
                    pass
            adjusted_width = (max_length + 1.5)
            column_letter = get_column_letter(column[0].column)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # 3. Format columns with thousands separators
        number_format = NamedStyle(name=f'number_format_{i}', number_format='### ### ### ##0')

        # Specify the columns to format based on float64 datatype
        float64_columns = pivot_table.select_dtypes(include=['float64']).columns
        print(float64_columns)
        for col in float64_columns:
            col_index = pivot_table.columns.get_loc(col) + 4
            print(col_index)  # 1-based index, starting from column 2
            col_letter = get_column_letter(col_index)
            print(col_letter)
            for cell in worksheet[col_letter]:
                # Start from the forth row assuming the first row is headers
                try:
                    if cell.value is not None:
                        formatted_value = "{:,.2f}".format(float(cell.value))
                        cell.value = float(cell.value)
                        cell.style = number_format
                except (ValueError, TypeError) as error:
                    print(error)

        # Add borders to all cells with data
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1,
                                       max_col=worksheet.max_column):
            for cell in row:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                     bottom=Side(style='thin'))

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for i, aksiya in enumerate(sorted(df_final['Aksiya'].unique()), start=1):
        # Filter the DataFrame for the current 'Aksiya' value
        df_aksiya = df_final[df_final['Aksiya'] == aksiya]

        # Perform the pivot operation
        pivoted_df = df_aksiya.pivot_table(index=['Region', 'ClientName', 'INN'],
                                           columns='DocName',
                                           values='OutQuantity',
                                           aggfunc='sum')
        # Calculate the GrandTotal
        pivoted_df['GrandTotal'] = pivoted_df.get('Оптовая реализация', 0) + pivoted_df.get('Товарная скидка', 0)
        doc_name1_total = pivoted_df['Оптовая реализация'].sum()
        doc_name2_total = pivoted_df['Товарная скидка'].sum()
        #
        total_row = pd.DataFrame({
            'Оптовая реализация': [doc_name1_total],
            'Товарная скидка': [doc_name2_total],
            'GrandTotal': [doc_name1_total + doc_name2_total]},
            index=pd.MultiIndex.from_tuples([('Grand Total', '', '')], names=['Region', 'ClientName', 'INN']))
        pivoted_df = pd.concat([pivoted_df, total_row], ignore_index=False)
        create_work_sheet(workbook=workbook, type_value=aksiya, pivot_table=pivoted_df, i=i)
    save_workbook(workbook=workbook, output_file_path=output_file_path)
