import os
from datetime import datetime

import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from sqlalchemy import create_engine

##################### LOADING IMPORTANT DATA ######################
# Load environment variables from the .env file
env_file_path = 'D:/Projects/.env'
load_dotenv(env_file_path)
# Giving output file name
output_file_path = 'HOURLY.xlsx'
##################### ACCESS ENV VARIABLES ######################
db_server = os.getenv("DB_SERVER")
db_database = os.getenv("DB_DATABASE_SERGELI")
db_user = os.getenv("DB_USER")
db_password = os.getenv("DB_PASSWORD")
db_port = os.getenv("DB_PORT")
db_driver_name = os.getenv("DB_DRIVER_NAME")

##################### CONNECTION STRING AND SQL QUERY ######################
# Construct the connection string
conn_str = f"mssql+pyodbc://{db_user}:{db_password}@{db_server}:{db_port}/{db_database}?driver={db_driver_name}"
engine = create_engine(conn_str)

sql_query = f"""
SELECT strdp.Name                                                               AS StoreDepName,
       g.Name                                                                   AS GoodName,
       a.Price,
       IIF(BasePrice = 0, 0, ROUND(100 * (a.Price - BasePrice) / BasePrice, 0)) AS MarkUp,
       il.ExpData                                                               AS ExpiryDate,
       P.FindName                                                               AS ProducerName,
       refpr.WholesalePrice                                                     as ReferencePrice
FROM PriceListLn a
         INNER JOIN
     IncomeLn il
     ON a.IncomeLnId = il.IncomeLnId
         INNER JOIN
     Income i ON il.IncomeId = i.IncomeId
         INNER JOIN
     Store st ON i.StoreId = st.StoreId
         INNER JOIN
     StoreDep strdp ON il.StoreDepId = strdp.StoreDepId
         INNER JOIN
     Good g ON il.GoodId = g.GoodId
         INNER JOIN
     Producer P ON g.ProducerId = P.ProducerId
         LEFT JOIN
     RefPrice refpr ON refpr.goodid = g.goodid
         LEFT JOIN
     GoodMaxPrice gMaxP ON il.GoodId = gMaxP.GoodId
WHERE il.Ostatok > 0
  AND il.Registr = 1
  and il.Ostatok - il.Reserved > 0
  AND i.StoreId = 1 -- Склад №1
ORDER BY st.Name,
         strdp.Name,
         g.Name,
         a.PriceListOrder;
"""
#####################  EXECUTION  ######################
price_df = pd.read_sql_query(sql_query, engine)
# Creating a new DataFrame and trim 'GoodName' and 'ProducerName' columns
price_df['GoodName'] = price_df['GoodName'].str.strip()
price_df['ProducerName'] = price_df['ProducerName'].str.strip()

# Filtering the ExpiryDate and making sure it is not an expired data
price_df['ExpiryDate'] = pd.to_datetime(price_df['ExpiryDate'], format='%d/%m/%Y')
today = datetime.now()
start_of_next_month = today.replace(day=1, month=today.month + 1).strftime('%Y-%m-%d')
price_df = price_df[price_df['ExpiryDate'] >= start_of_next_month]

# Concatenate 'GoodName', 'ProducerName', and 'Price' columns into 'TriadConcat'
price_df['TriadConcat'] = price_df['GoodName'] + price_df['ProducerName'] + price_df['Price'].astype(str)

# Drop duplicates based on 'TriadConcat' column
price_df.drop_duplicates(subset=['TriadConcat'], inplace=True)

# Drop TriadConcat because we don't need more.
price_df.drop(columns=['TriadConcat'], inplace=True)

# Fill NaN values in ReferencePrice Column with Price column
price_df.loc[:, 'ReferencePrice'] = price_df['ReferencePrice'].fillna(price_df['Price'])
price_df['ExpiryDate'] = price_df['ExpiryDate'].dt.strftime('%d/%m/%Y')

# Sort by GoodName for all StoreDepName excluding НБО
df_without_nbo = price_df[~price_df['StoreDepName'].str.contains('НБО')].sort_values(by='GoodName')
df_without_nbo.index = range(1, len(df_without_nbo) + 1)

# Sort by GoodName for only НБО StoreDepName
df_nbo = price_df[price_df['StoreDepName'].str.contains('НБО')].sort_values(by='GoodName')
df_nbo['MarkUp'] = 0
df_nbo.index = range(1, len(df_nbo) + 1)

# Concatenate both sorted dataframes
df_final = pd.concat([df_without_nbo, df_nbo], ignore_index=True)
df_final.index = range(1, len(df_final) + 1)
columns_export = ['GoodName', 'Price', 'MarkUp', 'ExpiryDate', 'ProducerName', 'ReferencePrice']
column_rename_dict = {
    'GoodName': 'Наименование медикаментов',
    'Price': 'Цена',
    'MarkUp': 'Наценка',
    'ExpiryDate': 'Срок годн.',
    'ProducerName': 'Производитель',
    'ReferencePrice': 'Референтная цена'
}

df_nbo = df_nbo[columns_export].rename(columns=column_rename_dict)
df_without_nbo = df_without_nbo[columns_export].rename(columns=column_rename_dict)
df_final = df_final[columns_export].rename(columns=column_rename_dict)

today = datetime.now().strftime('%d.%m.%Y')
file_names = {
    '1': {'name': f'ASKLEPIY_DISTRIBUTION_PRICE на {today}г.xlsx',
          'ReferencePrice': True,
          'data_frame': df_final},
    '2': {'name': f'ASKLEPIY_DISTRIBUTION_PRICE на {today}г_PDF.xlsx',
          'ReferencePrice': True,
          'data_frame': df_final},
    '3': {'name': f'Прайс-Лист {today}г.xlsx',
          'ReferencePrice': False,
          'data_frame': df_final},
    '4': {'name': f'Прайс-Лист {today} Без НБО.xlsx',
          'ReferencePrice': False,
          'data_frame': df_without_nbo},
    '5': {'name': f'Асклепий_Прайс_для_печати_{today}г.xlsx',
          'ReferencePrice': False,
          'data_frame': df_final},
}


def create_workbook(file_names) -> None:
    # Create a new workbook and select the active sheet
    for key, dict_value in file_names.items():
        print(f'Creating: {key}. {dict_value['name']}')
        wb = Workbook()
        wb.properties.creator = 'Behzod Xidirov'
        ws = wb.active
        ws.sheet_properties.tabColor = 'FFC4C4'
        ws.title = 'ASKLEPIY DISTRIBUTION'

        # Insert df into the worksheet
        for row_data in dataframe_to_rows(dict_value['data_frame'], header=True, index=True):
            ws.append(row_data)

        # Цена С НДС H1 COLUMN
        nds_formula = '=ROUND(C{}+(C{}*12%),0)'
        for row in range(3, len(dict_value['data_frame']) + 3):
            cell = ws[f'H{row}']
            cell.value = nds_formula.format(row, row)
            cell.number_format = '### ### ### ##0'
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="FFF455", end_color="FFF455", fill_type="solid")
            cell.font = Font(bold=True, italic=True)

        ws['A1'] = '№'
        ws['H1'] = 'Цена С НДС'
        # Adjust the width of the last column to fit the contents of cell H1
        last_column_width = len(str(ws['H1'].value)) + 2  # Add some padding
        ws.column_dimensions['H'].width = last_column_width

        # Colorize Reference Price G column
        for row in range(3, len(dict_value['data_frame']) + 3):
            cell = ws[f'G{row}']
            cell.fill = PatternFill(start_color="41B06E", end_color="FFF455", fill_type="solid")
            cell.font = Font(bold=True, italic=True)

        # Create border style
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        # some formatting to all cells
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border_style
                # Check if the cell contains a number
                if isinstance(cell.value, (int, float)):
                    # Apply space-separated formatting to numbers
                    cell.number_format = '### ### ### ##0;-### ### ### ##0'
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # Adjusting width all columns except the last one (column H)
        columns = list(ws.columns)
        for column in columns[:-1]:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                value = cell.value
                if value is not None:
                    max_length = max(max_length, len(str(value)))
            adjusted_width = (max_length + 2) * 1
            ws.column_dimensions[column_letter].width = adjusted_width

        # Adjust the width of the last column to fit the contents of cell H1
        last_column_width = len(str(ws['H1'].value)) + 2  # Add some padding
        ws.column_dimensions['H'].width = last_column_width

        # SOME COMMENTS AND CHANGES FOR BETTER QUALITY
        comment_markup = Comment("""ATTHD:
        Наценка от базовой цены
        """, "Behzod Xidirov")
        ws['D1'].comment = comment_markup
        # Save the changes to the workbook
        if not dict_value.get('ReferencePrice', False):
            ws.column_dimensions['G'].hidden = True

        wb.save(dict_value['name'])
    print('Successfully saved file!')


if __name__ == '__main__':
    create_workbook(file_names=file_names)
