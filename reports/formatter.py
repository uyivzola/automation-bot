import time

from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def formatter(df, output_file_path):
    print("Formatting🎨....")
    # Record the start time
    start_time = time.time()
    # Load the existing workbook
    workbook = load_workbook(output_file_path)

    # Access the default sheet (assuming it's the only sheet in the workbook)
    worksheet = workbook.active

    # 1. Color first row (headers) with 4CB9E7 color code
    header_style = NamedStyle(name='header_style',
                              fill=PatternFill(start_color='4CB9E7', end_color='4CB9E7', fill_type='solid'))

    for cell in worksheet[1]:
        cell.style = header_style

    # 2. Autofit all columns
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                value = str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)
            except:
                pass
        adjusted_width = (max_length + 1)
        column_letter = get_column_letter(column[0].column)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # 3. Format columns with thousands separators
    number_format = NamedStyle(name='number_format', number_format='### ### ### ##0')

    # Specify the columns to format based on float64 datatype
    float64_columns = df.select_dtypes(include=['float64']).columns

    for col in float64_columns:
        col_index = df.columns.get_loc(col) + 1  # 1-based index
        col_letter = get_column_letter(col_index)

        for cell in worksheet[col_letter][1:]:  # Start from the second row assuming the first row is headers
            try:
                formatted_value = "{:,.2f}".format(float(cell.value))
                cell.value = float(cell.value)
                cell.style = number_format
            except (ValueError, TypeError) as error:
                print(error)

    # Apply background color to all cells in the last column
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=worksheet.max_column,
                                   max_col=worksheet.max_column):
        for cell in row:
            cell.fill = PatternFill(start_color='A0D8B3', end_color='A0D8B3', fill_type='solid')

    # Add borders to all cells with data
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))

    workbook.save(output_file_path)
    end_time = time.time()

    # Calculate and print the elapsed time
    elapsed_time = round(end_time - start_time, 2)

    print(f"Formatting took: {elapsed_time} seconds.\n")
